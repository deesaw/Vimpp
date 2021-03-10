from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")

print(len(wb.sheetnames))
print(wb.sheetnames)

ws = wb['first sheet']

# read a cell
print(ws['A1'].value)
print(ws.cell(row=2,column=3).value)
# read a row
row = ws[3]
for cell in row:
    print(cell.value,end = " ")
# read a coll
col = ws['C']
for cell in col:
    print(cell.value)
# read a range
myrange = ws["A1:C4"]
for row in myrange:
    for cell in row:
        print(cell.value,end = " ")
    print()