from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "first sheet"
ws['A1']='Deloitte'
ws.cell(row=2,column=3,value="Consulting")
ws.append(["hello","how","are","you"])
ws.append([45,89,67,34,90])

ws1 = wb.create_sheet("second sheet")
ws1.append([1,2,3,4,5])

wb.save('sample.xlsx')